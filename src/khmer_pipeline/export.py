from __future__ import annotations
import codecs
import csv
import io
from collections import Counter
from datetime import datetime
from pathlib import Path
import openpyxl
from .models import PostprocessResult, ExportResult, Table, Cell
from .table_merge_pages import merge_document_tables

_XLSX_SHEET_NAME_ILLEGAL = set("[]:*?/\\")
_XLSX_SHEET_NAME_MAX_LEN = 31

# Excel requires a UTF-8 BOM prefix to open Khmer (or any non-ASCII) CSV text
# correctly; this is the str-level equivalent of encoding with "utf-8-sig".
_CSV_BOM = codecs.BOM_UTF8.decode("utf-8")  # U+FEFF, same char as the old literal "﻿"

_KHMER_TO_ARABIC: dict[str, str] = {
    "០": "0", "១": "1", "២": "2", "៣": "3", "៤": "4",
    "៥": "5", "៦": "6", "៧": "7", "៨": "8", "៩": "9",
}


def _convert_khmer_numerals(text: str) -> str:
    """Convert Khmer digit characters to Arabic digits.
    Applied to CSV cell text only when the user enables this option."""
    return "".join(_KHMER_TO_ARABIC.get(ch, ch) for ch in text)


def _convert_grid_numerals(grid: list[list[str]]) -> list[list[str]]:
    """Apply `_convert_khmer_numerals` to every cell in a row/col grid.
    Shared by `grid_to_csv` and `tables_to_xlsx` so both export formats
    convert numerals identically."""
    return [[_convert_khmer_numerals(cell) for cell in row] for row in grid]


def _validate_and_repair_table(table: Table) -> tuple[Table, bool]:
    """Pad rows that are shorter than the table's majority row length with
    empty placeholder cells, so the CSV/JSON grid is rectangular.

    Only pads short rows — rows longer than the majority length, or rows
    with col_ids outside range(target_cols), are left as-is. Assumes Surya's
    normal output of contiguous 0-indexed col_ids; degenerate cases are not
    fully normalized but will not raise."""
    cells = table.get("cells", [])
    if not cells:
        return table, False

    rows: dict[int, list[Cell]] = {}
    for c in cells:
        r = c.get("row_id", 0)
        rows.setdefault(r, []).append(c)

    row_lengths = [len(cols) for cols in rows.values()]
    if len(set(row_lengths)) == 1:
        return table, False  # already consistent

    target_cols = Counter(row_lengths).most_common(1)[0][0]
    repaired_cells = list(cells)
    for row_idx, row_cells in rows.items():
        existing_col_ids = {c.get("col_id", 0) for c in row_cells}
        for col_id in range(target_cols):
            if col_id not in existing_col_ids:
                repaired_cells.append({
                    "row_id": row_idx,
                    "col_id": col_id,
                    "text_lines": [],
                    "bbox": [],
                })

    repaired_table = dict(table)
    repaired_table["cells"] = repaired_cells
    repaired_table["was_repaired"] = True
    return repaired_table, True


def export(result: PostprocessResult, convert_numerals: bool = False, repair_tables: bool = False,
           stitch_pages: bool = False) -> ExportResult:
    """Build the final export payload from a `PostprocessResult`: the document JSON plus
    one CSV per table (optionally repairing ragged rows and/or stitching continuation
    tables across pages first). Returns an `ExportResult`."""
    # Repair tables in place before building the JSON, so was_repaired and
    # the padded cell grid are reflected in both document_json and the CSVs.
    # This mutates the input PostprocessResult's page.tables; export() is the
    # final pipeline stage, so nothing reads the pre-repair state afterward.
    # (This also relies on CorrectedPageResult.tables being the same list
    # object as SuryaPageResult.tables, per postprocess.py's _correct_page,
    # so app.py's was_repaired badge sees the repair too.)
    # Repair is opt-in: analysts must explicitly request it, since it
    # rewrites the detected cell grid.
    if repair_tables:
        for page in result.pages:
            for t_idx, table in enumerate(page.tables):
                page.tables[t_idx], _ = _validate_and_repair_table(table)

    document_json = _build_document_json(result)

    if stitch_pages:
        # Join continuation tables across pages into one logical table per section;
        # the merged tables become the CSV output an analyst gets.
        merged = merge_document_tables(result.pages)
        document_json["document_tables"] = _build_merged_tables_json(result.source_name, merged)
        tables_csv = [(_make_doc_table_id(result.source_name, i), _table_to_csv(t, convert_numerals))
                      for i, t in enumerate(merged)]
    else:
        tables_csv = []
        for page in result.pages:
            for t_idx, table in enumerate(page.tables):
                table_id = _make_table_id(result.source_name, page.page_index, t_idx)
                tables_csv.append((table_id, _table_to_csv(table, convert_numerals)))

    return ExportResult(
        source_name=result.source_name,
        document_json=document_json,
        tables_csv=tables_csv,
    )


def _make_doc_table_id(source_name: str, n: int) -> str:
    return f"{Path(source_name).stem}_table{n + 1}"


def _build_merged_tables_json(source_name: str, merged: list[Table]) -> list[dict]:
    out = []
    for i, table in enumerate(merged):
        out.append({
            "table_id": _make_doc_table_id(source_name, i),
            "source_pages": table.get("source_pages", []),
            "rows": table["rows"],
            "cols": table["cols"],
            "cells": [
                {
                    "row": c.get("row_id", 0),
                    "col": c.get("col_id") or 0,
                    "text": " ".join(
                        t["text"] for t in (c.get("text_lines") or []) if t.get("text")
                    ).strip(),
                }
                for c in table.get("cells", [])
            ],
        })
    return out


def _make_table_id(source_name: str, page_index: int, table_index: int) -> str:
    return f"{Path(source_name).stem}_page{page_index + 1}_table{table_index + 1}"


def grid_to_csv(grid: list[list[str]], convert_numerals: bool = False) -> str:
    """Render a row/col grid of strings as CSV text with a UTF-8 BOM prefix (required
    for Excel to open Khmer text correctly). Optionally converts Khmer digits to
    Arabic first."""
    buf = io.StringIO()
    buf.write(_CSV_BOM)
    writer = csv.writer(buf)
    if not grid:
        return buf.getvalue()
    if convert_numerals:
        grid = _convert_grid_numerals(grid)
    writer.writerows(grid)
    return buf.getvalue()


def _sanitize_sheet_name(table_id: str, used: set[str]) -> str:
    name = "".join(ch for ch in table_id if ch not in _XLSX_SHEET_NAME_ILLEGAL)
    name = name[:_XLSX_SHEET_NAME_MAX_LEN] or "Sheet"
    if name not in used:
        return name
    # De-dupe by appending _2, _3, ... while staying within the length limit.
    n = 2
    while True:
        suffix = f"_{n}"
        candidate = name[:_XLSX_SHEET_NAME_MAX_LEN - len(suffix)] + suffix
        if candidate not in used:
            return candidate
        n += 1


def tables_to_xlsx(tables: list[tuple[str, list[list[str]]]], convert_numerals: bool = False) -> bytes:
    """Build an XLSX workbook with one sheet per non-blank table, named from each
    table's id (sanitized/de-duped for Excel's sheet-name rules). Returns the
    workbook's bytes."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # default sheet; only keep sheets for non-blank tables
    used_names: set[str] = set()
    for table_id, grid in tables:
        if not grid or not any(any(cell for cell in row) for row in grid):
            continue
        if convert_numerals:
            grid = _convert_grid_numerals(grid)
        sheet_name = _sanitize_sheet_name(table_id, used_names)
        used_names.add(sheet_name)
        ws = wb.create_sheet(sheet_name)
        for row in grid:
            ws.append(row)

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _table_to_csv(table: Table, convert_numerals: bool = False) -> str:
    cells = table.get("cells", [])
    if not cells:
        return grid_to_csv([])
    max_row = max(c.get("row_id", 0) for c in cells) + 1
    max_col = max((c.get("col_id") or 0) for c in cells) + 1
    grid = [[""] * max_col for _ in range(max_row)]
    for c in cells:
        r = c.get("row_id", 0)
        col = c.get("col_id") or 0
        text = " ".join(
            t["text"] for t in (c.get("text_lines") or []) if t.get("text")
        ).strip()
        if 0 <= r < max_row and 0 <= col < max_col:
            grid[r][col] = text
    return grid_to_csv(grid, convert_numerals)


def _build_document_json(result: PostprocessResult) -> dict:
    return {
        "source_name": result.source_name,
        "extracted_at": datetime.utcnow().isoformat() + "Z",
        "page_count": len(result.pages),
        "pages": [
            {
                "page_index": page.page_index,
                "qwen_used": page.qwen_used,
                "corrected_text": page.corrected_text,
                "tables": [
                    {
                        "table_index": t_idx,
                        "table_id": _make_table_id(result.source_name, page.page_index, t_idx),
                        "was_repaired": table.get("was_repaired", False),
                        "rows": table["rows"],
                        "cols": table["cols"],
                        "cells": [
                            {
                                "row": c.get("row_id", 0),
                                "col": c.get("col_id") or 0,
                                "text": " ".join(
                                    t["text"] for t in (c.get("text_lines") or []) if t.get("text")
                                ).strip(),
                            }
                            for c in table.get("cells", [])
                        ],
                    }
                    for t_idx, table in enumerate(page.tables)
                ],
            }
            for page in result.pages
        ],
    }
